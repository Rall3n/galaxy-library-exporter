import argparse
import json
import re
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path


platform_names = {
    'amazon': 'Prime Gaming (Amazon)',
    'epic': 'Epic Games',
    'gog': 'GOG Galaxy',
    'origin': 'Origin',
    'psn': 'PlayStation',
    'rockstar': 'Rockstar Games',
    'steam': 'Steam',
    'twitch': 'Prime Gaming (Twitch)',
    'uplay': 'UPlay',
    'battlenet': 'Battle.Net'
}


def format_achievements_value(game):
    achievements = game['myAchievementsCount']

    if achievements['all'] is not None and achievements['all'] > 0:
        return f"{achievements['unlocked']} / {achievements['all']}"


class Galaxy2LibraryExporter:
    args = None
    db_connection = None
    db_cursor = None

    columns = [
        {
            'title': 'Title',
            'field': 'title',
            'width': 75,
            'filter': True
        }
    ]

    def __init__(self):
        self.__setup_arguments()

    def __setup_arguments(self):
        arg_parser = argparse.ArgumentParser()

        arg_parser.add_argument('db_path', action='store')
        arg_parser.add_argument('-o', '--output', action='store', default='galaxy-library', dest='output_filename', help='Filename of the output file')
        arg_parser.add_argument('-t', '--tags', action='store', nargs='*', default=False, dest='tags', metavar="Tag", help='Which game tags to include')
        arg_parser.add_argument('-p', '--platforms', nargs='+', dest='platforms', metavar="Platform", help='Which platforms to include. All by default.')

        self.args = arg_parser.parse_args()

    def __setup_db(self):
        self.db_connection = sqlite3.connect(self.args.db_path)
        self.db_cursor = self.db_connection.cursor()

    def __close_db(self):
        self.db_cursor.close()
        self.db_connection.close()

    def _read_games_from_db(self):
        # Get game piece types
        gamePieceTypes = {}

        self.db_cursor.execute('SELECT * FROM GamePieceTypes')
        for pieceType in self.db_cursor.fetchall():
            gamePieceTypes[pieceType[0]] = pieceType[1]

        # Get games data
        games = {}
        dlcs = list()

        self.db_cursor.execute(f'SELECT releaseKey, gamePieceTypeId, value FROM GamePieces WHERE releaseKey IN (SELECT releaseKey FROM GameLinks) ORDER BY releaseKey ASC, gamePieceTypeId ASC;')

        for game_id, pieceTypeId, value in self.db_cursor.fetchall():
            pieceType = gamePieceTypes[pieceTypeId]

            value = json.loads(value)
            if len(value.keys()) == 1:
                value = list(value.values())[0]

            if game_id not in games.keys():
                games[game_id] = dict()

            games[game_id][pieceType] = value

            if pieceType == 'dlcs':
                dlcs.extend(value)

        if isinstance(self.args.tags, list):
            self.columns.append({
                'title': 'Tags',
                'field': 'tags',
                'filter': True
            })

            query = 'SELECT releaseKey, group_concat(tag) FROM UserReleaseTags'
            query += f' WHERE tag IN {repr(self.args.tags).replace("[","(").replace("]",")")}' if len(self.args.tags) > 0 else ''
            query += ' GROUP BY releaseKey'

            self.db_cursor.execute(query)

            for game_id, tags in self.db_cursor.fetchall():
                if game_id not in games.keys():
                    continue

                games[game_id]['tags'] = tags

        return games, dlcs

    def run(self):
        self.__setup_db()

        # Get games data
        games, dlcs = self._read_games_from_db()

        self.__close_db()

        # Sort games to platforms
        platforms = {}

        for game_id, game in games.items():
            platform = game_id.split('_')[0]

            # Generic are games added manually. They have no value to us
            if platform == 'generic':
                continue
            
            if self.args.platforms and platform not in self.args.platforms:
                continue

            if platform not in platforms.keys():
                platforms[platform] = []

            # We do not need dlcs in the list (for now).
            # A dlc is either marked by being a dlc entry in a game,
            # or by having a value on the `parent` piece type
            if game_id in dlcs or game.get('parent', False):
                continue

            platforms[platform].append(game)

        wb = Workbook()
        wb.remove(wb.active)

        self.columns.append({
            'title': 'Achievements',
            'field': 'myAchievementsCount',
            'formatValue': format_achievements_value
        })

        # Keep filterable columns at the front, makes setting up auto_filter easier
        self.columns.sort(key=lambda x: not x.get('filter', False))

        for platform in platforms.keys():
            platforms[platform].sort(key=lambda game: game['title'].lower())

        for p_indx, platform in enumerate(platforms.keys()):
            ws = wb.create_sheet(platform_names.get(platform, platform), p_indx)
            filters = []

            ws.title = platform_names.get(platform, platform)

            for c_indx, column in enumerate(self.columns):
                LETTER = get_column_letter(c_indx+1)

                ws.column_dimensions[LETTER].width = column.get('width', 25)
                ws[f'{LETTER}1'] = column['title']
                ws[f'{LETTER}1'].font = Font(bold=True)

                if column.get('filter', False):
                    filters.append(LETTER)

            game_count = 0

            for g_indx, game in enumerate(platforms[platform]):
                game_count += 1
                ROW = g_indx+2

                for c_indx, column in enumerate(self.columns):
                    LETTER = get_column_letter(c_indx+1)

                    if column['field'] in game.keys():
                        if column.get('formatValue', False):
                            ws[f'{LETTER}{ROW}'] = column['formatValue'](game)
                        elif game[column['field']]:
                            ws[f'{LETTER}{ROW}'] = game[column['field']]

            ws.auto_filter.ref = f'{filters[0]}1:{filters[-1]}{game_count+1}'

        wb.save(f'{self.args.output_filename}.xlsx')

if __name__ == '__main__':
    cmd = Galaxy2LibraryExporter()
    cmd.run()
